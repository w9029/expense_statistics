from __future__ import annotations

import argparse
import json
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from urllib.parse import urlencode
from typing import Any

import requests
from openpyxl import load_workbook

DEFAULT_API_BASE_URL = "http://127.0.0.1:8080/api/v1"
# DEFAULT_SOURCE_PATH = Path(__file__).resolve().parent / "importdata.xlsx"
DEFAULT_SOURCE_PATH = "F:\\OneDrive\\记账.xlsx"
DEFAULT_SHEET_NAME = "日常开销"
DEFAULT_MERGE_CATEGORY_NAME = "合并消费"
EXPECTED_HEADERS = ("日期", "消费数额", "消费名称", "消费种类", "币种", "备注")
_CURRENCY_ALIASES = {
    "": "JPY",
    "JPY": "JPY",
    "CNY": "CNY",
    "USD": "USD",
    "日元": "JPY",
    "人民币": "CNY",
    "美元": "USD",
}


class ApiError(RuntimeError):
    def __init__(self, message: str, *, status_code: int | None = None, code: str | None = None) -> None:
        super().__init__(message)
        self.status_code = status_code
        self.code = code


@dataclass(frozen=True)
class CategoryRecord:
    id: str
    name: str
    is_merge_category: bool


@dataclass(frozen=True)
class SourceRow:
    excel_row: int
    spent_at: str | None
    amount: str
    name: str
    category_name: str | None
    currency: str
    currency_explicit: bool
    note: str | None


@dataclass(frozen=True)
class ChildExpense:
    excel_row: int
    amount: str
    name: str
    category_name: str
    note: str | None


@dataclass(frozen=True)
class NormalExpenseEntry:
    excel_row: int
    spent_at: str
    amount: str
    name: str
    category_name: str
    currency: str
    note: str | None


@dataclass(frozen=True)
class MergedExpenseEntry:
    excel_row: int
    spent_at: str
    total_amount: str
    name: str
    currency: str
    note: str | None
    children: list[ChildExpense]


@dataclass(frozen=True)
class ValidationIssue:
    excel_row: int
    message: str


@dataclass(frozen=True)
class CheckResult:
    categories_by_name: dict[str, CategoryRecord]
    merge_category: CategoryRecord
    entries: list[NormalExpenseEntry | MergedExpenseEntry]
    normal_count: int
    merged_count: int
    merged_child_count: int
    daily_total_mismatches: list["DailyTotalMismatch"]
    importable_entries: list[NormalExpenseEntry | MergedExpenseEntry]


@dataclass(frozen=True)
class ExistingExpenseSummary:
    spent_at: str
    name: str
    original_amount: str
    original_currency: str
    expense_type: str


@dataclass(frozen=True)
class DailyTotalMismatch:
    spent_at: str
    source_totals: dict[str, str]
    existing_totals: dict[str, str]
    status: str
    unmatched_existing_entries: list[str]


@dataclass(frozen=True)
class ComparableExpense:
    spent_at: str
    name: str
    amount: str
    currency: str
    expense_type: str


class ExpenseApiClient:
    def __init__(
        self,
        *,
        api_base_url: str,
        token: str | None,
        email: str | None,
        password: str | None,
        timeout_seconds: float,
    ) -> None:
        self.api_base_url = api_base_url.rstrip("/")
        self.token = token.strip() if token else None
        self.email = email.strip() if email else None
        self.password = password if password else None
        self.timeout_seconds = timeout_seconds
        self.session = requests.Session()

    def list_categories(self, account_book_id: str) -> list[CategoryRecord]:
        payload = self._request_json(
            "GET",
            f"/account-books/{account_book_id}/expense-categories",
        )
        if not isinstance(payload, list):
            raise ApiError("category list response is not an array")
        categories: list[CategoryRecord] = []
        for item in payload:
            categories.append(
                CategoryRecord(
                    id=str(item["id"]),
                    name=str(item["name"]),
                    is_merge_category=bool(item["is_merge_category"]),
                )
            )
        return categories

    def create_normal_expense(self, account_book_id: str, payload: dict[str, Any]) -> dict[str, Any]:
        return self._request_json(
            "POST",
            f"/account-books/{account_book_id}/expenses/normal",
            body=payload,
        )

    def create_merged_expense(self, account_book_id: str, payload: dict[str, Any]) -> dict[str, Any]:
        return self._request_json(
            "POST",
            f"/account-books/{account_book_id}/expenses/merged",
            body=payload,
        )

    def list_all_expenses(
        self,
        account_book_id: str,
        *,
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> list[ExistingExpenseSummary]:
        page = 1
        page_size = 100
        items: list[ExistingExpenseSummary] = []

        while True:
            query_params: dict[str, str | int] = {"page": page, "page_size": page_size}
            if date_from:
                query_params["date_from"] = date_from
            if date_to:
                query_params["date_to"] = date_to
            query = urlencode(query_params)
            payload = self._request_json(
                "GET",
                f"/account-books/{account_book_id}/expenses?{query}",
            )
            batch = payload.get("items")
            if not isinstance(batch, list):
                raise ApiError("expense list response is missing items")

            for item in batch:
                items.append(
                    ExistingExpenseSummary(
                        spent_at=str(item["spent_at"]),
                        name=str(item["name"]),
                        original_amount=str(item["original_amount"]),
                        original_currency=str(item["original_currency"]),
                        expense_type=str(item.get("expense_type", "")),
                    )
                )

            total = int(payload.get("total", 0))
            if page * page_size >= total or not batch:
                break
            page += 1

        return items

    def _request_json(self, method: str, path: str, body: dict[str, Any] | None = None) -> Any:
        self._ensure_token()
        response = self.session.request(
            method=method,
            url=f"{self.api_base_url}{path}",
            headers=self._headers(),
            json=body,
            timeout=self.timeout_seconds,
        )
        if response.status_code == 401 and self.email and self.password:
            self._login(force=True)
            response = self.session.request(
                method=method,
                url=f"{self.api_base_url}{path}",
                headers=self._headers(),
                json=body,
                timeout=self.timeout_seconds,
            )
        return self._unwrap_response(response)

    def _ensure_token(self) -> None:
        if self.token:
            return
        if not self.email or not self.password:
            raise ApiError("token is required, or provide --email and --password for auto login")
        self._login(force=False)

    def _login(self, *, force: bool) -> None:
        if self.token and not force:
            return
        if not self.email or not self.password:
            raise ApiError("cannot login without --email and --password")
        response = self.session.post(
            f"{self.api_base_url}/identity/login",
            headers={"Accept": "application/json", "Content-Type": "application/json"},
            json={"email": self.email, "password": self.password},
            timeout=self.timeout_seconds,
        )
        payload = self._unwrap_response(response)
        token = payload.get("access_token")
        if not token:
            raise ApiError("login response does not include access_token")
        self.token = str(token)

    def _headers(self) -> dict[str, str]:
        headers = {"Accept": "application/json"}
        if self.token:
            headers["Authorization"] = f"Bearer {self.token}"
        return headers

    @staticmethod
    def _unwrap_response(response: requests.Response) -> Any:
        try:
            payload = response.json()
        except json.JSONDecodeError as exc:
            raise ApiError(
                f"request failed with non-JSON response: {response.status_code}",
                status_code=response.status_code,
            ) from exc

        if response.ok and payload.get("ok") is True:
            return payload.get("data")

        message = payload.get("message") or f"request failed: {response.status_code}"
        code = payload.get("error")
        raise ApiError(message, status_code=response.status_code, code=code)


def add_common_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument(
        "--api-base-url",
        default=DEFAULT_API_BASE_URL,
        help="Backend API base URL. Default: http://127.0.0.1:8080/api/v1",
    )
    parser.add_argument(
        "--account-book-id",
        required=True,
        help="Target account book ID.",
    )
    parser.add_argument(
        "--token",
        help="Access token for backend API.",
    )
    parser.add_argument(
        "--email",
        help="Optional login email used when --token is omitted or expires mid-run.",
    )
    parser.add_argument(
        "--password",
        help="Optional login password used when --token is omitted or expires mid-run.",
    )
    parser.add_argument(
        "--source",
        default=str(DEFAULT_SOURCE_PATH),
        help="Path to importdata.xlsx. Default: documents/database/expense_data_import/importdata.xlsx",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET_NAME,
        help="Worksheet name. Default: 日常开销",
    )
    parser.add_argument(
        "--merge-category-name",
        default=DEFAULT_MERGE_CATEGORY_NAME,
        help="Exact category name used for merged expense parents. Default: 合并消费",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=20.0,
        help="HTTP timeout in seconds. Default: 20.",
    )
    parser.add_argument(
        "--lookback-days",
        type=int,
        default=0,
        help="Only process Excel entries in the last N natural days. 0 means no limit.",
    )


def run_check(
    *,
    api_base_url: str,
    account_book_id: str,
    token: str | None,
    email: str | None,
    password: str | None,
    source_path: str | Path,
    sheet_name: str,
    merge_category_name: str,
    timeout_seconds: float,
    lookback_days: int,
) -> CheckResult:
    client = ExpenseApiClient(
        api_base_url=api_base_url,
        token=token,
        email=email,
        password=password,
        timeout_seconds=timeout_seconds,
    )
    categories = client.list_categories(account_book_id)
    categories_by_name, merge_category = validate_category_catalog(categories, merge_category_name)
    rows, issues = load_source_rows(Path(source_path), sheet_name)
    entries, entry_issues = build_entries(rows, categories_by_name)
    entries = filter_entries_by_lookback(entries, lookback_days)
    issues.extend(entry_issues)
    if issues:
        raise ValidationFailed(issues)
    normal_count = sum(isinstance(entry, NormalExpenseEntry) for entry in entries)
    merged_count = sum(isinstance(entry, MergedExpenseEntry) for entry in entries)
    merged_child_count = sum(
        len(entry.children) for entry in entries if isinstance(entry, MergedExpenseEntry)
    )
    date_from, date_to = derive_entry_date_range(entries)
    existing_expenses = client.list_all_expenses(
        account_book_id,
        date_from=date_from,
        date_to=date_to,
    ) if entries else []
    daily_total_mismatches = compare_daily_totals(entries, existing_expenses)
    importable_entries = select_importable_entries(entries, existing_expenses, daily_total_mismatches)
    return CheckResult(
        categories_by_name=categories_by_name,
        merge_category=merge_category,
        entries=entries,
        normal_count=normal_count,
        merged_count=merged_count,
        merged_child_count=merged_child_count,
        daily_total_mismatches=daily_total_mismatches,
        importable_entries=importable_entries,
    )


class ValidationFailed(RuntimeError):
    def __init__(self, issues: list[ValidationIssue]) -> None:
        super().__init__(f"validation failed with {len(issues)} issue(s)")
        self.issues = issues


def validate_category_catalog(
    categories: list[CategoryRecord],
    merge_category_name: str,
) -> tuple[dict[str, CategoryRecord], CategoryRecord]:
    categories_by_name: dict[str, CategoryRecord] = {}
    for category in categories:
        if category.name in categories_by_name:
            raise ApiError(f"duplicate category name returned by backend: {category.name}")
        categories_by_name[category.name] = category

    merge_category = categories_by_name.get(merge_category_name)
    if merge_category is None:
        raise ApiError(f'merge category "{merge_category_name}" does not exist in the target account book')
    if not merge_category.is_merge_category:
        raise ApiError(f'category "{merge_category_name}" exists but is not a merge category')
    return categories_by_name, merge_category


def load_source_rows(source_path: Path, sheet_name: str) -> tuple[list[SourceRow], list[ValidationIssue]]:
    if not source_path.exists():
        raise FileNotFoundError(f"source workbook not found: {source_path}")

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f'worksheet "{sheet_name}" not found. Available sheets: {available}')

    sheet = workbook[sheet_name]
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=6, values_only=True))
    headers = tuple("" if cell is None else str(cell).strip() for cell in header_cells)
    if headers != EXPECTED_HEADERS:
        raise ValueError(f"worksheet headers must be exactly {EXPECTED_HEADERS}, got {headers}")

    rows: list[SourceRow] = []
    issues: list[ValidationIssue] = []
    for excel_row, values in enumerate(
        sheet.iter_rows(min_row=2, min_col=1, max_col=6, values_only=True),
        start=2,
    ):
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        try:
            spent_at = parse_excel_date(values[0], excel_row)
            amount = parse_excel_amount(values[1], excel_row)
            name = parse_required_text(values[2], excel_row, "消费名称")
            category_name = parse_optional_text(values[3])
            currency, currency_explicit = parse_currency(values[4], excel_row)
            note = parse_optional_text(values[5])
        except ValueError as exc:
            issues.append(ValidationIssue(excel_row=excel_row, message=str(exc)))
            continue
        rows.append(
            SourceRow(
                excel_row=excel_row,
                spent_at=spent_at,
                amount=amount,
                name=name,
                category_name=category_name,
                currency=currency,
                currency_explicit=currency_explicit,
                note=note,
            )
        )
    return rows, issues


def build_entries(
    rows: list[SourceRow],
    categories_by_name: dict[str, CategoryRecord],
) -> tuple[list[NormalExpenseEntry | MergedExpenseEntry], list[ValidationIssue]]:
    entries: list[NormalExpenseEntry | MergedExpenseEntry] = []
    issues: list[ValidationIssue] = []

    index = 0
    while index < len(rows):
        row = rows[index]
        if row.spent_at is None:
            issues.append(
                ValidationIssue(
                    excel_row=row.excel_row,
                    message="child row cannot appear before a parent row with date",
                )
            )
            index += 1
            continue

        if row.category_name:
            category = categories_by_name.get(row.category_name)
            if category is None:
                issues.append(
                    ValidationIssue(
                        excel_row=row.excel_row,
                        message=f'category "{row.category_name}" does not exist in target account book',
                    )
                )
            elif category.is_merge_category:
                issues.append(
                    ValidationIssue(
                        excel_row=row.excel_row,
                        message=f'normal expense row cannot use merge category "{row.category_name}"',
                    )
                )
            else:
                entries.append(
                    NormalExpenseEntry(
                        excel_row=row.excel_row,
                        spent_at=row.spent_at,
                        amount=row.amount,
                        name=row.name,
                        category_name=row.category_name,
                        currency=row.currency,
                        note=row.note,
                    )
                )
            index += 1
            continue

        child_rows: list[SourceRow] = []
        lookahead = index + 1
        while lookahead < len(rows) and rows[lookahead].spent_at is None:
            child_rows.append(rows[lookahead])
            lookahead += 1
        if not child_rows:
            issues.append(
                ValidationIssue(
                    excel_row=row.excel_row,
                    message="top-level row with blank category must be followed by at least one child row",
                )
            )
            index += 1
            continue

        children: list[ChildExpense] = []
        for child in child_rows:
            if not child.category_name:
                issues.append(
                    ValidationIssue(
                        excel_row=child.excel_row,
                        message="merged child row must have a category name",
                    )
                )
                continue
            category = categories_by_name.get(child.category_name)
            if category is None:
                issues.append(
                    ValidationIssue(
                        excel_row=child.excel_row,
                        message=f'category "{child.category_name}" does not exist in target account book',
                    )
                )
                continue
            if category.is_merge_category:
                issues.append(
                    ValidationIssue(
                        excel_row=child.excel_row,
                        message=f'merged child row cannot use merge category "{child.category_name}"',
                    )
                )
                continue
            if child.currency_explicit and child.currency != row.currency:
                issues.append(
                    ValidationIssue(
                        excel_row=child.excel_row,
                        message=(
                            "merged child row currency must be blank/default or equal to parent currency "
                            f"({row.currency}), got {child.currency}"
                        ),
                    )
                )
                continue
            children.append(
                ChildExpense(
                    excel_row=child.excel_row,
                    amount=child.amount,
                    name=child.name,
                    category_name=child.category_name,
                    note=child.note,
                )
            )

        if len(children) == len(child_rows):
            entries.append(
                MergedExpenseEntry(
                    excel_row=row.excel_row,
                    spent_at=row.spent_at,
                    total_amount=row.amount,
                    name=row.name,
                    currency=row.currency,
                    note=row.note,
                    children=children,
                )
            )
        index = lookahead

    return entries, issues


def parse_excel_date(value: Any, excel_row: int) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if text == "":
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
        raise ValueError("日期 must be Excel date or YYYY-MM-DD/YYYY/MM/DD string")


def parse_excel_amount(value: Any, excel_row: int) -> str:
    if value is None or str(value).strip() == "":
        raise ValueError("消费数额 is required")
    try:
        amount = Decimal(str(value).strip())
    except InvalidOperation as exc:
        raise ValueError("消费数额 must be numeric") from exc
    if amount <= 0:
        raise ValueError("消费数额 must be greater than 0")
    if amount.as_tuple().exponent < -2:
        raise ValueError("消费数额 must have at most 2 decimal places")
    normalized = amount.quantize(Decimal("0.01"))
    return f"{normalized:.2f}"


def parse_required_text(value: Any, excel_row: int, column_name: str) -> str:
    text = parse_optional_text(value)
    if not text:
        raise ValueError(f"{column_name} is required")
    return text


def parse_optional_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_currency(value: Any, excel_row: int) -> tuple[str, bool]:
    raw = parse_optional_text(value) or ""
    normalized = _CURRENCY_ALIASES.get(raw)
    if normalized is None:
        raise ValueError(
            f'币种 "{raw}" is unsupported. Allowed values are blank, 日元, 人民币, 美元, JPY, CNY, USD'
        )
    return normalized, raw != ""


def print_check_summary(result: CheckResult) -> None:
    print("Check passed.")
    print(f"Normal expenses : {result.normal_count}")
    print(f"Merged expenses : {result.merged_count}")
    print(f"Merged children : {result.merged_child_count}")
    print(f"Total API writes: {result.normal_count + result.merged_count}")
    if not result.daily_total_mismatches:
        print("Daily total compare: matched")
        importable_dates = sorted({entry.spent_at for entry in result.importable_entries})
        print(f"Importable entries: {len(result.importable_entries)}")
        print(f"Importable dates: {importable_dates if importable_dates else '[]'}")
        return

    print(f"Daily total compare: {len(result.daily_total_mismatches)} mismatched day(s)")
    supplementable = [m for m in result.daily_total_mismatches if m.status == "supplementable"]
    conflicts = [m for m in result.daily_total_mismatches if m.status == "conflict"]

    print(f"\nSupplementable mismatches: {len(supplementable)}")
    for mismatch in supplementable:
        print(
            f"  {mismatch.spent_at}: "
            f"source={format_currency_totals(mismatch.source_totals)} "
            f"existing={format_currency_totals(mismatch.existing_totals)}"
        )

    print(f"\nConflict mismatches: {len(conflicts)}")
    for mismatch in conflicts:
        print(
            f"  {mismatch.spent_at}: "
            f"source={format_currency_totals(mismatch.source_totals)} "
            f"existing={format_currency_totals(mismatch.existing_totals)}"
        )
        if mismatch.unmatched_existing_entries:
            print(f"    unmatched_existing={mismatch.unmatched_existing_entries}")
    importable_dates = sorted({entry.spent_at for entry in result.importable_entries})
    print(f"\nImportable entries: {len(result.importable_entries)}")
    print(f"Importable dates: {importable_dates if importable_dates else '[]'}")


def print_validation_failed(exc: ValidationFailed) -> None:
    print(f"Check failed with {len(exc.issues)} issue(s):")
    for issue in sorted(exc.issues, key=lambda item: item.excel_row):
        print(f"  row {issue.excel_row}: {issue.message}")


def compare_daily_totals(
    entries: list[NormalExpenseEntry | MergedExpenseEntry],
    existing_expenses: list[ExistingExpenseSummary],
) -> list[DailyTotalMismatch]:
    source_daily = aggregate_source_daily_totals(entries)
    existing_daily = aggregate_existing_daily_totals(existing_expenses)
    source_records = aggregate_source_daily_records(entries)
    existing_records = aggregate_existing_daily_records(existing_expenses)

    mismatches: list[DailyTotalMismatch] = []
    all_dates = sorted(set(source_daily.keys()) | set(existing_daily.keys()))
    for spent_at in all_dates:
        source_totals = normalize_total_map(source_daily.get(spent_at, {}))
        existing_totals = normalize_total_map(existing_daily.get(spent_at, {}))
        if source_totals != existing_totals:
            unmatched_existing_entries = find_unmatched_existing_entries(
                source_records.get(spent_at, []),
                existing_records.get(spent_at, []),
            )
            status = "supplementable" if not unmatched_existing_entries else "conflict"
            mismatches.append(
                DailyTotalMismatch(
                    spent_at=spent_at,
                    source_totals=source_totals,
                    existing_totals=existing_totals,
                    status=status,
                    unmatched_existing_entries=unmatched_existing_entries,
                )
            )
    return mismatches


def select_importable_entries(
    entries: list[NormalExpenseEntry | MergedExpenseEntry],
    existing_expenses: list[ExistingExpenseSummary],
    daily_total_mismatches: list[DailyTotalMismatch],
) -> list[NormalExpenseEntry | MergedExpenseEntry]:
    supplementable_dates = {
        mismatch.spent_at
        for mismatch in daily_total_mismatches
        if mismatch.status == "supplementable"
    }
    if not supplementable_dates:
        return []

    existing_records = aggregate_existing_daily_records(existing_expenses)
    remaining_by_date: dict[str, dict[tuple[str, str, str, str], int]] = {}
    for spent_at in supplementable_dates:
        counters: dict[tuple[str, str, str, str], int] = defaultdict(int)
        for record in existing_records.get(spent_at, []):
            counters[(record.expense_type, record.name, record.currency, record.amount)] += 1
        remaining_by_date[spent_at] = counters

    importable: list[NormalExpenseEntry | MergedExpenseEntry] = []
    for entry in entries:
        spent_at = entry.spent_at
        if spent_at not in supplementable_dates:
            continue

        comparable = comparable_from_entry(entry)
        key = (comparable.expense_type, comparable.name, comparable.currency, comparable.amount)
        remaining = remaining_by_date[spent_at]
        if remaining.get(key, 0) > 0:
            remaining[key] -= 1
            continue
        importable.append(entry)

    return importable


def aggregate_source_daily_totals(
    entries: list[NormalExpenseEntry | MergedExpenseEntry],
) -> dict[str, dict[str, Decimal]]:
    daily: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(lambda: Decimal("0.00")))
    for entry in entries:
        if isinstance(entry, NormalExpenseEntry):
            daily[entry.spent_at][entry.currency] += Decimal(entry.amount)
            continue
        daily[entry.spent_at][entry.currency] += Decimal(entry.total_amount)
    return daily


def aggregate_existing_daily_totals(
    expenses: list[ExistingExpenseSummary],
) -> dict[str, dict[str, Decimal]]:
    daily: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(lambda: Decimal("0.00")))
    for expense in expenses:
        if expense.expense_type == "merged_child":
            continue
        daily[expense.spent_at][expense.original_currency] += Decimal(expense.original_amount)
    return daily


def aggregate_source_daily_records(
    entries: list[NormalExpenseEntry | MergedExpenseEntry],
) -> dict[str, list[ComparableExpense]]:
    daily: dict[str, list[ComparableExpense]] = defaultdict(list)
    for entry in entries:
        if isinstance(entry, NormalExpenseEntry):
            daily[entry.spent_at].append(
                ComparableExpense(
                    spent_at=entry.spent_at,
                    name=entry.name,
                    amount=entry.amount,
                    currency=entry.currency,
                    expense_type="normal",
                )
            )
            continue
        daily[entry.spent_at].append(
            ComparableExpense(
                spent_at=entry.spent_at,
                name=entry.name,
                amount=entry.total_amount,
                currency=entry.currency,
                expense_type="merged_parent",
            )
        )
    return daily


def aggregate_existing_daily_records(
    expenses: list[ExistingExpenseSummary],
) -> dict[str, list[ComparableExpense]]:
    daily: dict[str, list[ComparableExpense]] = defaultdict(list)
    for expense in expenses:
        if expense.expense_type == "merged_child":
            continue
        daily[expense.spent_at].append(
            ComparableExpense(
                spent_at=expense.spent_at,
                name=expense.name,
                amount=expense.original_amount,
                currency=expense.original_currency,
                expense_type=expense.expense_type,
            )
        )
    return daily


def find_unmatched_existing_entries(
    source_records: list[ComparableExpense],
    existing_records: list[ComparableExpense],
) -> list[str]:
    remaining: dict[tuple[str, str, str, str], int] = defaultdict(int)
    for record in source_records:
        remaining[(record.expense_type, record.name, record.currency, record.amount)] += 1

    unmatched: list[str] = []
    for record in existing_records:
        key = (record.expense_type, record.name, record.currency, record.amount)
        if remaining.get(key, 0) > 0:
            remaining[key] -= 1
            continue
        unmatched.append(format_comparable_expense(record))
    return unmatched


def normalize_total_map(totals: dict[str, Decimal]) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for currency, amount in sorted(totals.items()):
        quantized = amount.quantize(Decimal("0.01"))
        if quantized == Decimal("0.00"):
            continue
        normalized[currency] = f"{quantized:.2f}"
    return normalized


def format_currency_totals(totals: dict[str, str]) -> str:
    if not totals:
        return "{}"
    parts = [f"{currency}={amount}" for currency, amount in totals.items()]
    return "{%s}" % ", ".join(parts)


def format_comparable_expense(record: ComparableExpense) -> str:
    return (
        f"{record.expense_type}:{record.name}:{record.currency}:{record.amount}"
    )


def comparable_from_entry(entry: NormalExpenseEntry | MergedExpenseEntry) -> ComparableExpense:
    if isinstance(entry, NormalExpenseEntry):
        return ComparableExpense(
            spent_at=entry.spent_at,
            name=entry.name,
            amount=entry.amount,
            currency=entry.currency,
            expense_type="normal",
        )
    return ComparableExpense(
        spent_at=entry.spent_at,
        name=entry.name,
        amount=entry.total_amount,
        currency=entry.currency,
        expense_type="merged_parent",
    )


def filter_entries_by_lookback(
    entries: list[NormalExpenseEntry | MergedExpenseEntry],
    lookback_days: int,
) -> list[NormalExpenseEntry | MergedExpenseEntry]:
    if lookback_days <= 0:
        return entries

    cutoff = date.today() - timedelta(days=lookback_days - 1)
    filtered: list[NormalExpenseEntry | MergedExpenseEntry] = []
    for entry in entries:
        spent_at = datetime.strptime(entry.spent_at, "%Y-%m-%d").date()
        if spent_at >= cutoff:
            filtered.append(entry)
    return filtered


def derive_entry_date_range(
    entries: list[NormalExpenseEntry | MergedExpenseEntry],
) -> tuple[str | None, str | None]:
    if not entries:
        return None, None
    dates = sorted(entry.spent_at for entry in entries)
    return dates[0], dates[-1]
